from PageObjects.HomePage import HomePage
from PageObjects.MobilesListPage import SearchListPage
from Utility.BaseClass import BaseClass
from openpyxl import Workbook


class TestFlipKartWebsite(BaseClass, HomePage, SearchListPage):

    def test_retrieve(self):
        book = Workbook()
        sheet = book.active

        self.fn_explicitWait(self.login_X_button)

        self.driver.find_element_by_xpath(self.login_X_button).click()

        self.driver.find_element_by_xpath(self.search_input_text).send_keys("iphone")

        self.driver.find_element_by_xpath(self.search_button).click()

        self.fn_explicitWait(self.low_to_high)

        self.driver.find_element_by_xpath(self.low_to_high).click()

        all_mobiles = self.driver.find_elements_by_xpath(self.list_all_mobiles)

        for i in range(1, len(all_mobiles)-1):
            self.fn_explicitWait("(//div[@class='_1HmYoV _35HD7C']/div[@class='bhgxx2 col-12-12']/div/div/div/a/div[2]/div[1]/div[1])[{}]".format(i))
            mobile_name = self.driver.find_element_by_xpath("(//div[@class='_1HmYoV _35HD7C']/div[@class='bhgxx2 col-12-12']/div/div/div/a/div[2]/div[1]/div[1])[{}]".format(i)).text
            mobile_prize = self.driver.find_element_by_xpath("((//div[@class='_1HmYoV _35HD7C']/div[@class='bhgxx2 col-12-12'])[{}]/div/div/div/a/div[2]/div[2]/div/div/div)[1]".format(i)).text
            mobile_rating = self.driver.find_element_by_xpath("((//div[@class='_1HmYoV _35HD7C']/div[@class='bhgxx2 col-12-12'])[{}]/div/div/div/a/div[2]/div/div[2]/span[2]/span/span)[1]".format(i)).text
            mo_pr = mobile_prize[1:]
            prize_int = int(mo_pr.replace(",", ""))

            if prize_int > 40000:
                break

            print(mobile_name, end=" ")
            print(prize_int, end=" ")
            print(mobile_rating.replace(" Ratings", ""))
            sheet.cell(i, 1, mobile_name)
            sheet.cell(i, 2, mobile_prize)
            sheet.cell(i, 3, mobile_rating.replace(" Ratings", ""))

        book.save("C:\\Users\\User\\Documents\\Siddappa\\Guru\\Python\\pythonAssignment\\Reports\\Output.xlsx")
